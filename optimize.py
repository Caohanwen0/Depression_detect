# -*- coding:utf-8 -*-
# 1. 使用预训练模型RoBERTa-wwm-ext-large
# 2. 使用warmup策略
# 3. set param.requires_grad = True
from sklearn.model_selection import train_test_split
from transformers import BertTokenizer
from transformers import BertModel
from transformers import AutoModel
import torch
from torch import nn
import numpy as np
import pandas as pd
from transformers import BertForSequenceClassification, AdamW, BertConfig, get_linear_schedule_with_warmup
from torch.utils.data import TensorDataset, DataLoader, RandomSampler, SequentialSampler
import openpyxl
from sklearn.utils import class_weight
import sys
from sklearn.metrics import classification_report

ENABLE_GPU = True
PADDING_LENGTH = 512

print('\nDefining hyperparameters\n')
epochs = 30
batch_size = int(sys.argv[1])
learning_rate = int(sys.argv[2])
grad_accumulate = int(sys.argv[3])

path = 'saved_models/without_blank_lr='+str(learning_rate)+"bs="+str(batch_size)+"g_a="+str(grad_accumulate)+'.pt' 


if ENABLE_GPU:
    device = torch.device("cuda")

wb = openpyxl.load_workbook('data/data_clean_all.xlsx')
ws = wb.active
serial = []
label = []
for col in ws['A']:
    serial.append(col.value) 
for col in ws['L']:
    label.append(col.value)
del serial[0]
del label[0]
texts = []
print("  \nReading files to construct raw dataset\n   ")
for serial_num in serial:
    f = open('data/text_pool/' + str(serial_num) + ".txt", 'r')
    raw = f.read()
    texts.append(raw)
    f.close()

print("delete blank data")
i = 0
while i < len(texts):
    if len(texts[i].strip()) == 0:
        del texts[i]
        del label[i]
    i += 1

print("\nSplit dataset\n")

train_text, temp_text, train_labels, temp_labels = train_test_split(
    texts, 
    label, 
    random_state = 2022, 
    test_size=0.2, 
    stratify = label)
val_text, test_text, val_labels, test_labels = train_test_split(
    temp_text, 
    temp_labels,
    random_state= 2022, 
    test_size=0.5, 
    stratify=temp_labels)

print("The length of train, val and test is{}, {}, {}".format(len(train_text), len(val_text), len(test_text)))

print('\ncompute the class weights\n')
class_weights = class_weight.compute_class_weight('balanced', classes = np.unique(train_labels), y = train_labels)
# print("Class Weights:",class_weights)
# Class Weights: [0.74559778 1.51792453]
# 抑郁人数:非抑郁人数～1:2


# define bert model
tokenizer = BertTokenizer.from_pretrained('hfl/chinese-roberta-wwm-ext-large')

# tokenize and encode sequences in the training set
tokens_train = tokenizer.batch_encode_plus(
    train_text,
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)

# tokenize and encode sequences in the validation set
tokens_val = tokenizer.batch_encode_plus(
    val_text,
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)

# tokenize and encode sequences in the test set
tokens_test = tokenizer.batch_encode_plus(
    test_text,
    max_length = PADDING_LENGTH,
    padding='max_length',
    truncation=True
)

print('\nconvert lists to tensors\n')

train_seq = torch.tensor(tokens_train['input_ids'])
train_mask = torch.tensor(tokens_train['attention_mask'])
train_y = torch.tensor(train_labels)

val_seq = torch.tensor(tokens_val['input_ids'])
val_mask = torch.tensor(tokens_val['attention_mask'])
val_y = torch.tensor(val_labels)

test_seq = torch.tensor(tokens_test['input_ids'])
test_mask = torch.tensor(tokens_test['attention_mask'])
test_y = torch.tensor(test_labels)

train_data = TensorDataset(train_seq, train_mask, train_y)
train_sampler = RandomSampler(train_data)
train_dataloader = DataLoader(train_data, sampler=train_sampler, batch_size=batch_size)
val_data = TensorDataset(val_seq, val_mask, val_y)
val_sampler = SequentialSampler(val_data)
val_dataloader = DataLoader(val_data, sampler = val_sampler, batch_size=batch_size)

# # unfreeze
# for param in bert.parameters():
#     param.requires_grad = True

# model definition
class BERT_Arch(nn.Module):
    def __init__(self):
        super(BERT_Arch, self).__init__()
        self.bert = BertModel.from_pretrained('hfl/chinese-roberta-wwm-ext-large')
        self.fc2 = nn.Linear(1024 ,2) # binary classification
        self.softmax = nn.LogSoftmax(dim=1)

    #define the forward pass
    def forward(self, sent_id, mask):
        _, cls_hs = self.bert(sent_id, attention_mask=mask, return_dict=False)
        x = self.fc2(cls_hs)
        x = self.softmax(x)
        return x

print("\nLoad model\n")
model = BERT_Arch()
# model.load_state_dict(torch.load('saved_models/without_blank_lr=1e-05bs=8g_a=4.pt', map_location = "cpu"))
if ENABLE_GPU:
    model = model.to(device)

optimizer = AdamW(model.parameters(), lr = learning_rate)   
weights= torch.tensor(class_weights,dtype=torch.float)
if ENABLE_GPU:  
    weights = weights.to(device)
cross_entropy  = nn.NLLLoss(weight=weights)

print("len(train_data)",len(train_data))
total_step = (len(train_dataloader)) * epochs
warm_up_ratio = 0

# scheduler = get_linear_schedule_with_warmup(optimizer,
#                                             num_warmup_steps= 0,
#                                             num_training_steps=total_step)



def train():
    print('\n...Training model...')
    model.train()
    print("...Training finish...\n")
    total_loss = 0
    # empty list to save model predictions
    total_preds=[]
    # iterate over batches
    # correct = 0
    for step,batch in enumerate(train_dataloader):
    # progress update after every 50 batches.
        if step % 50 == 0 and not step == 0:
            print('Batch {:>5,}  of  {:>5,}.'.format(step, len(train_dataloader)))
        # push the batch to gpu
        if ENABLE_GPU:
            batch = [r.to(device) for r in batch]
        sent_id, mask, labels = batch
        model.zero_grad()        
        preds = model(sent_id, mask)
        loss = cross_entropy(preds, labels) # compute loss
        total_loss = total_loss + loss.item()
        loss = loss / grad_accumulate 
        loss.backward() # backward propagation
        torch.nn.utils.clip_grad_norm_(model.parameters(), 1.0)
        if ((step + 1) % grad_accumulate == 0) or (step == (len(train_dataloader) - 1)):
            optimizer.step() # update model
            # scheduler.step() # update learning rate
            optimizer.zero_grad() # clear gradient to zero
        # preds=preds.detach().cpu().numpy()
        # total_preds.append(preds)
    # compute the training loss of the epoch
    avg_loss = total_loss / len(train_dataloader)
    # total_preds  = np.concatenate(total_preds, axis=0)
    return avg_loss, total_preds

# function for evaluating the model
def evaluate():
    print("\nEvaluating...\n")
    # deactivate dropout layers
    model.eval() # disable the dropout layer for evaluating
    total_loss, total_accuracy = 0, 0
    # empty list to save the model predictions
    total_preds = []
    for step,batch in enumerate(val_dataloader):
        if step % 50 == 0 and not step == 0:
            print('Batch {:>5,}  of  {:>5,}.'.format(step, len(val_dataloader)))
        if ENABLE_GPU:
            batch = [t.to(device) for t in batch]
        sent_id, mask, labels = batch
        with torch.no_grad():
            preds = model(sent_id, mask) # make predictions
            preds = preds.detach().cpu().numpy() ##
            preds = np.argmax(preds, axis = 1) ##
            labels = np.array(labels.tolist())
            acc = np.sum(preds == labels) / len(labels) #?????
            total_accuracy = total_accuracy + acc
    # compute the validation loss of the epoch
    avg_acc = total_accuracy / len(val_dataloader)
    # total_preds  = np.concatenate(total_preds, axis=0)
    return avg_acc, total_preds

def fine_tunning():
    # set initial acc to zero
    best_valid_acc = 0
    # empty lists to store training and validation loss of each epoch
    train_losses=[]
    valid_acc_list =[] # store validation accuracy
    saving_cnt = 0 # count storing model times 
    #for each epoch
    for epoch in range(epochs):
        
        print('\n Epoch {:} / {:}'.format(epoch + 1, epochs))
        #train model
        train_loss, _ = train()
        
        #evaluate model
        valid_acc, _ = evaluate()
        #save the best model
        if valid_acc > best_valid_acc:
            best_valid_acc = valid_acc
            print("\nsaving the {}.th model, with lr = {}, bs = {}, grad_accumulate ={}\n".format(saving_cnt, learning_rate, batch_size, grad_accumulate))
            torch.save(model.state_dict(), path) # save the newest model
            saving_cnt += 1
        train_losses.append(train_loss)
        valid_acc_list.append(valid_acc)

        print(f'\nTraining Loss: {train_loss:.3f}')
        print(f'Validation Accuracy: {valid_acc:.3f}')
    # 将valid_acc_list保存下来,用于日后画图使用
    df = pd.DataFrame(valid_acc_list)
    csv_name = 'saved_models/large_without_blank_lr='+str(learning_rate)+"bs="+str(batch_size)+"g_a="+str(grad_accumulate)+".csv" 
    df.to_csv(csv_name)

def check_performance():
    preds_list = np.array([])
    for i in range(0,10):
        with torch.no_grad():
            if ENABLE_GPU:
                preds = model(test_seq[i*44:i*44+44].to(device), test_mask[i*44:i*44+44].to(device))
            else:
                preds = model(test_seq[i*44:i*44+44], test_mask[i*44:i*44+44])
        preds = preds.detach().cpu().numpy()
        preds = np.argmax(preds, axis = 1)
        preds_list = np.concatenate([preds_list, preds])
    with open('saved_models/result.txt', 'a') as f:
        f.write("path={}".format(path))
        f.write("unfreezed large model(delete blank data):lr = {}, bs = {}, grad_accumulate = {}\n".format(learning_rate, batch_size, grad_accumulate))
        f.write(classification_report(test_y, preds_list))
        f.write('\n\n\n')
    # 将验证集上的结果保存到analysis.csv中
    # data = {'text': test_text, 'labels':test_labels, 'preds' :preds_list}
    # df = pd.DataFrame(data)
    # df.to_csv('saved_models/analysis_aftertrained.csv', sep = ',', mode = 'w')



if __name__ == "__main__":
    fine_tunning()
    check_performance()

